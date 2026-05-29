import streamlit as st
from openpyxl import load_workbook
import csv
import tempfile
import os

st.title("Large Excel Column Extractor")

uploaded_file = st.file_uploader(
    "Upload Excel File",
    type=["xlsx"]
)

if uploaded_file:

    tmp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_xlsx.write(uploaded_file.getbuffer())
    tmp_xlsx.close()

    with st.spinner("Reading header..."):

        wb = load_workbook(
            tmp_xlsx.name,
            read_only=True,
            data_only=True
        )

        ws = wb.active

        headers = [c.value for c in next(ws.iter_rows(max_row=1))]

    selected_cols = st.multiselect(
        "Select output columns",
        headers
    )

    if st.button("Extract") and selected_cols:

        idx = [
            headers.index(col)
            for col in selected_cols
        ]

        out_csv = tempfile.NamedTemporaryFile(
            delete=False,
            suffix=".csv",
            mode="w",
            newline="",
            encoding="utf-8"
        )

        writer = csv.writer(out_csv)

        writer.writerow(selected_cols)

        total = ws.max_row or 1
        progress = st.progress(0)

        count = 0

        for row in ws.iter_rows(min_row=2, values_only=True):

            writer.writerow([
                row[i] if i < len(row) else ""
                for i in idx
            ])

            count += 1

            if count % 5000 == 0:
                progress.progress(min(count / total, 1.0))

        out_csv.close()
        wb.close()

        with open(out_csv.name, "rb") as f:

            st.download_button(
                "Download Output CSV",
                f,
                file_name="filtered_output.csv",
                mime="text/csv"
            )

        os.unlink(tmp_xlsx.name)
